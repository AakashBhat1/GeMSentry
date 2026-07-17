#!/usr/bin/env python3
"""
Import ETSPL tender master sheet → history.json (BE-12).

Reads `TENDER MASTER SHEET(ETSPL) 2025- 26.xlsx` sheet
`(TENDER DETAILS (PARTICIPATED)` and writes `history.json`.

Also prints suggested buyer_affinity + value_preference priors derived
from history (median-based). Does NOT auto-train ML weights.

Idempotent: re-running overwrites history.json from the xlsx source of truth.
Never modifies the xlsx.
"""
from __future__ import annotations

import json
import os
import re
import statistics
import sys
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# Repo root = parent of tools/
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_XLSX = os.path.join(ROOT, "TENDER MASTER SHEET(ETSPL) 2025- 26.xlsx")
DEFAULT_OUT = os.path.join(ROOT, "history.json")
SHEET_NAME = "(TENDER DETAILS (PARTICIPATED)"

# Keyword → business_line id (mirrors company profile seeds)
_LINE_KEYWORDS = {
    "drone": [
        "drone", "drones", "uav", "unmanned", "gis", "mapping", "surveillance",
        "reconnaissance", "aerostat"
    ],
    "power_supply": [
        "power supply", "ac-dc", "ac dc", "rectifier", "alternator", "ups",
        "voltage", "lvpsu", "hvpsu", "power unit", "convertor", "converter",
        "battery charger", "psu", "amplifier"
    ],
    "ai_it": [
        "artificial intelligence", "ai based", "ai-based", "software", "server",
        "radar", "cctv", "camera", "laptop", "notebook", "electronics",
        "display", "network", "router"
    ],
}


def _parse_value(raw):
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return int(raw)
    s = str(raw).strip()
    if not s or s.upper() in ("N/A", "NA", "-", "—"):
        return None
    # Take first number-like token (handles multi-part / GST text)
    s = s.replace(",", "")
    m = re.search(r'(\d+(?:\.\d+)?)', s)
    if not m:
        return None
    try:
        return int(float(m.group(1)))
    except ValueError:
        return None


def _parse_won(result_raw):
    """Return True/False/None from RESULT WON/LOST cell."""
    if result_raw is None or str(result_raw).strip() == "":
        return None
    s = str(result_raw).upper()
    if "WON" in s or re.search(r'\bL\s*-?\s*1\b', s) or "RANK - L1" in s or "RANK L1" in s:
        # L1 alone is weak; require WON or explicit RANK - L1
        if "WON" in s or "RANK - L1" in s or "RANK L1" in s or re.search(r'WON\s*L\s*-?\s*1', s):
            return True
    if "LOST" in s:
        return False
    if re.search(r'L\s*-?\s*[2-9]', s) or re.search(r'RANK\s*-?\s*L\s*[2-9]', s):
        return False
    return None


def _guess_business_line(description: str):
    if not description:
        return None
    low = description.lower()
    best_id = None
    best_hits = 0
    for line_id, kws in _LINE_KEYWORDS.items():
        hits = sum(1 for kw in kws if kw in low)
        if hits > best_hits:
            best_hits = hits
            best_id = line_id
    return best_id if best_hits > 0 else None


def _normalize_buyer(org: str) -> str:
    if not org:
        return "UNKNOWN"
    s = re.sub(r'\s+', ' ', str(org)).strip().upper()
    # Collapse common variants
    aliases = {
        "INDIAN AIR FORCE": ["IAF", "INDIAN AIRFORCE"],
        "INDIAN ARMY": ["ARMY"],
        "INDIAN NAVY": ["NAVY"],
        "BHARAT PETROLEUM": ["BPCL", "BHARAT PETROLEUM CORPORATION"],
    }
    for canon, alts in aliases.items():
        if s == canon or any(a in s for a in alts) or canon in s:
            return canon
    return s


def import_history(xlsx_path=None, out_path=None):
    xlsx_path = xlsx_path or DEFAULT_XLSX
    out_path = out_path or DEFAULT_OUT

    if not os.path.exists(xlsx_path):
        print(f"ERROR: xlsx not found: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET_NAME]

    # Header row is row 2
    headers = {}
    max_col = ws.max_column or 46
    max_row = ws.max_row or 100
    for c in range(1, max_col + 1):
        h = ws.cell(2, c).value
        if h:
            headers[str(h).strip().replace("\n", " ")] = c

    def col(*names):
        for n in names:
            if n in headers:
                return headers[n]
        # fuzzy
        for key, idx in headers.items():
            for n in names:
                if n.lower() in key.lower():
                    return idx
        return None

    c_org = col("ORGANISATION")
    c_loc = col("LOCATIOIN/SITE", "LOCATION/SITE")
    c_desc = col("DESCRIPTION")
    c_cat = col("WORK CATEGORY")
    c_val = col("TENDER VALUE")
    c_result = col("RESULT WON/LOST", "RESULT\nWON/LOST")
    c_status = col("STATUS")
    c_tid = col("TENDER ID")

    rows = []
    for r in range(3, max_row + 1):
        org = ws.cell(r, c_org).value if c_org else None
        desc = ws.cell(r, c_desc).value if c_desc else None
        # Skip empty
        if org is None and desc is None:
            continue
        # Skip header-like repeats
        if str(org).strip().upper() == "ORGANISATION":
            continue

        value = _parse_value(ws.cell(r, c_val).value if c_val else None)
        result_raw = ws.cell(r, c_result).value if c_result else None
        won = _parse_won(result_raw)
        buyer = _normalize_buyer(org)
        category = ws.cell(r, c_cat).value if c_cat else None
        location = ws.cell(r, c_loc).value if c_loc else None
        tender_id = ws.cell(r, c_tid).value if c_tid else None
        status = ws.cell(r, c_status).value if c_status else None
        bl = _guess_business_line(str(desc or ""))

        rows.append({
            "buyer": buyer,
            "value_inr": value,
            "category": str(category).strip() if category else None,
            "result": str(result_raw).strip() if result_raw else None,
            "business_line": bl,
            "won": won,
            "description": str(desc).strip() if desc else None,
            "location": str(location).strip() if location else None,
            "tender_id": str(tender_id).strip() if tender_id else None,
            "status": str(status).strip() if status else None,
        })

    wb.close()

    payload = {
        "version": 1,
        "source": os.path.basename(xlsx_path),
        "sheet": SHEET_NAME,
        "row_count": len(rows),
        "rows": rows,
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"Wrote {len(rows)} rows -> {out_path}")

    # --- Suggested buyer_affinity (count-weighted, max 1.0) ---
    counts = Counter(r["buyer"] for r in rows if r.get("buyer"))
    max_c = max(counts.values()) if counts else 1
    affinity = {}
    for buyer, c in counts.most_common():
        # floor 0.3, scale to 1.0 by relative frequency
        affinity[buyer] = round(0.3 + 0.7 * (c / max_c), 2)

    print("\n=== Suggested buyer_affinity (paste into company_profile.json) ===")
    print(json.dumps(affinity, indent=2))

    # --- Suggested value_preference from numeric values ---
    values = sorted(r["value_inr"] for r in rows if isinstance(r.get("value_inr"), int) and r["value_inr"] > 0)
    if values:
        med = int(statistics.median(values))
        # sweet band: ~25th–90th percentile-ish around median
        p25 = values[max(0, int(len(values) * 0.25) - 1)]
        p90 = values[min(len(values) - 1, int(len(values) * 0.90))]
        # clamp to reasonable defaults floor
        sweet_min = max(100000, int(p25 * 0.5))
        sweet_max = max(sweet_min + 1, int(p90))
        print("\n=== Suggested value_preference ===")
        print(json.dumps({
            "sweet_min_inr": sweet_min,
            "sweet_max_inr": sweet_max,
            "median_value_inr": med,
            "n_values": len(values),
        }, indent=2))
    else:
        print("\nNo numeric tender values found for value_preference suggestion.")

    # Label stats
    labeled = sum(1 for r in rows if r.get("won") is not None)
    print(f"\nLabeled won/lost rows: {labeled}/{len(rows)} (ML weight-learning deferred).")
    return payload


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else None
    out = sys.argv[2] if len(sys.argv) > 2 else None
    import_history(xlsx, out)
