"""
Export a workspace's tender metadata to a decision-ready Excel workbook.

Repeatable replacement for the one-off complete_summary generator:
  - Overview sheet (counts, generated-at, scoring fingerprint)
  - Pursue / Review / Drop tabs sorted by priority (best first)
  - All_Tenders master sheet
  - New_Since_Last tab — tenders added since the previous export
    (tracked in complete_summary/.export_state.json)

Columns include priority/fit/risk, value, days left (red <5 / amber <10),
EMD, exemptions, and dual hyperlinks (GeM portal + local PDF).

Usage:
  python tools/export_summary.py                     # main workspace
  python tools/export_summary.py --workspace personel
  python tools/export_summary.py --output my.xlsx
"""
import argparse
import datetime
import json
import os
import sys
import urllib.parse

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

import paths  # noqa: E402
import scraper  # noqa: E402

HEADERS = [
    ("Priority", 9), ("Fit", 6), ("Risk", 6), ("Recommendation", 14),
    ("Status", 14), ("Bid Number", 20), ("Title", 60), ("Business Line", 18),
    ("Buyer / Department", 40), ("Est. Value (INR)", 15), ("Days Left", 9),
    ("End Date", 18), ("EMD", 22), ("Startup Exemption", 16),
    ("MSE Exemption", 16), ("Confidence", 10), ("Keyword", 22),
    ("GeM Portal", 11), ("Local PDF", 10),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill("solid", fgColor="F8CBAD")
AMBER_FILL = PatternFill("solid", fgColor="FFE699")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
LINK_FONT = Font(color="0563C1", underline="single")


def days_left(tender):
    end = scraper.parse_gem_date(tender.get("end_date"))
    if not end:
        return None
    return round((end - datetime.datetime.now()).total_seconds() / 86400.0, 1)


def file_uri(rel_path):
    if not rel_path:
        return None
    abs_path = rel_path if os.path.isabs(rel_path) else os.path.join(paths.ROOT, rel_path)
    if not os.path.exists(abs_path):
        return None
    return "file:///" + urllib.parse.quote(abs_path.replace("\\", "/"), safe="/:")


def sort_key(tender):
    pr = (tender.get("analysis") or {}).get("priority_score")
    return pr if pr is not None else -1


def write_sheet(wb, name, tenders, tab_color=None):
    ws = wb.create_sheet(name)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    for col, (title, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    for row, t in enumerate(sorted(tenders, key=sort_key, reverse=True), 2):
        a = t.get("analysis") or {}
        dl = days_left(t)
        rec = a.get("recommendation")
        bl = a.get("business_line") or {}
        values = [
            a.get("priority_score"), a.get("fit_score"), a.get("score"),
            rec or "—", t.get("status"), t.get("bid_no"),
            (t.get("title") or "")[:200], bl.get("label") or "—",
            (t.get("department") or a.get("buyer_org") or "")[:120],
            a.get("est_value_inr"), dl, t.get("end_date") or "—",
            a.get("emd_status") or "—", a.get("startup_exemption") or "—",
            a.get("mse_exemption") or "—", a.get("confidence"),
            t.get("keyword") or "", None, None,
        ]
        for col, v in enumerate(values, 1):
            ws.cell(row=row, column=col, value=v)

        if rec == "Pursue":
            ws.cell(row=row, column=4).fill = GREEN_FILL
        if dl is not None and dl <= 10:
            ws.cell(row=row, column=11).fill = RED_FILL if dl <= 5 else AMBER_FILL
        if a.get("est_value_inr") is not None:
            ws.cell(row=row, column=10).number_format = "#,##0"

        gem = ws.cell(row=row, column=18)
        if t.get("pdf_url"):
            gem.value = "Open Bid"
            gem.hyperlink = t["pdf_url"]
            gem.font = LINK_FONT
        else:
            gem.value = "—"
        pdf = ws.cell(row=row, column=19)
        uri = file_uri(t.get("local_pdf_path"))
        if uri:
            pdf.value = "Open PDF"
            pdf.hyperlink = uri
            pdf.font = LINK_FONT
        else:
            pdf.value = "—"
    return ws


def write_overview(wb, tenders, fingerprint, new_count):
    ws = wb.create_sheet("Overview", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    title = ws.cell(row=1, column=1, value="GeM Tender Summary")
    title.font = Font(size=16, bold=True)
    rows = [
        ("Generated", datetime.datetime.now().strftime("%d %b %Y, %I:%M %p")),
        ("Total tenders", len(tenders)),
        ("New since last export", new_count),
        ("Scoring fingerprint", fingerprint or "—"),
        ("", ""),
    ]
    recs, statuses = {}, {}
    for t in tenders:
        r = (t.get("analysis") or {}).get("recommendation") or "Unanalyzed"
        recs[r] = recs.get(r, 0) + 1
        s = t.get("status") or "—"
        statuses[s] = statuses.get(s, 0) + 1
    rows += [("By recommendation", "")] + sorted(recs.items())
    rows += [("", ""), ("By status", "")] + sorted(statuses.items())
    for i, (k, v) in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=(v == ""))
        ws.cell(row=i, column=2, value=v)
    return ws


def export_workbook(tenders_dir, output=None):
    """
    Export a workspace's metadata to the CENTRAL reports folder:
    tenders/reports/tender_summary_<profile>.xlsx — one folder, one clearly
    named workbook per profile (main, personel, ...). Lives outside every
    downloads/ tree so clearing a workspace never destroys workbooks.
    Returns a summary dict, or None when the workspace has no tenders.
    """
    label = scraper.workspace_label(tenders_dir)
    reports_dir = os.path.join(paths.TENDERS_DIR, "reports")
    os.makedirs(reports_dir, exist_ok=True)
    out_path = output or os.path.join(reports_dir, f"tender_summary_{label}.xlsx")
    state_path = os.path.join(reports_dir, f".export_state_{label}.json")

    tenders = list(scraper.load_existing_metadata(tenders_dir).values())
    if not tenders:
        return None

    prev_bids = set()
    if os.path.exists(state_path):
        try:
            with open(state_path, encoding="utf-8") as f:
                prev_bids = set(json.load(f).get("bids", []))
        except Exception:
            prev_bids = set()
    new_tenders = [t for t in tenders if t.get("bid_no") not in prev_bids]

    cfg = scraper.load_scoring_config()
    profile = scraper.load_company_profile()
    fingerprint = scraper.scoring_fingerprint(cfg, profile)

    by_rec = {"Pursue": [], "Review": [], "Drop": []}
    for t in tenders:
        rec = (t.get("analysis") or {}).get("recommendation")
        if rec in by_rec:
            by_rec[rec].append(t)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_overview(wb, tenders, fingerprint, len(new_tenders) if prev_bids else 0)
    write_sheet(wb, "Pursue", by_rec["Pursue"], tab_color="2E7D32")
    write_sheet(wb, "Review", by_rec["Review"], tab_color="F9A825")
    write_sheet(wb, "Drop", by_rec["Drop"], tab_color="C62828")
    if prev_bids and new_tenders:
        write_sheet(wb, "New_Since_Last", new_tenders, tab_color="1565C0")
    write_sheet(wb, "All_Tenders", tenders)
    wb.save(out_path)

    with open(state_path, "w", encoding="utf-8") as f:
        json.dump({
            "bids": sorted(t.get("bid_no") for t in tenders if t.get("bid_no")),
            "exported_at": datetime.datetime.now().isoformat(timespec="seconds"),
        }, f, indent=2)

    return {
        "output": out_path,
        "total": len(tenders),
        "pursue": len(by_rec["Pursue"]),
        "review": len(by_rec["Review"]),
        "drop": len(by_rec["Drop"]),
        "new": len(new_tenders) if prev_bids else None,
    }


def main():
    parser = argparse.ArgumentParser(description="Export tender metadata to Excel")
    parser.add_argument("--workspace", default="main",
                        help="'main' or a subfolder of tenders/ (default: main)")
    parser.add_argument("--output", help="Output .xlsx path (default: complete_summary/)")
    args = parser.parse_args()

    if args.workspace == "main":
        tenders_dir = paths.TENDERS_DIR
    else:
        tenders_dir = os.path.join(paths.TENDERS_DIR, args.workspace)

    summary = export_workbook(tenders_dir, output=args.output)
    if summary is None:
        sys.exit("No tenders in metadata; nothing to export.")
    new_note = summary["new"] if summary["new"] is not None else "(first tracked export)"
    print(f"Exported {summary['total']} tenders -> {paths.repo_relative(summary['output'])}")
    print(f"  Pursue: {summary['pursue']}, Review: {summary['review']}, "
          f"Drop: {summary['drop']}, New since last: {new_note}")


if __name__ == "__main__":
    main()
