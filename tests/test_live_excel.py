import os
import sys
import time
import shutil
import datetime
import tempfile
import openpyxl
import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from gemsentry.live_excel import LiveExcelManager, build_curated_workbook



@pytest.fixture
def temp_manager(tmp_path):
    # Set up a manager with a very short timeout (e.g. 0.5 seconds) for testing
    mgr = LiveExcelManager(timeout_seconds=1)
    # Redirect directories to tmp_path
    mgr.reports_dir = str(tmp_path / "reports")
    mgr.live_dir = str(tmp_path / "reports" / "live")
    mgr.daily_dir = str(tmp_path / "reports" / "daily_exports")
    os.makedirs(mgr.live_dir, exist_ok=True)
    os.makedirs(mgr.daily_dir, exist_ok=True)
    mgr.live_excel_path = os.path.join(mgr.live_dir, "live_working_excel.xlsx")
    mgr.state_file = os.path.join(mgr.live_dir, ".live_session.json")
    mgr.is_active = False
    mgr.status = "idle"
    mgr.tender_bids = []
    mgr.update_count = 0
    mgr.created_at = None
    mgr.last_updated_at = None
    yield mgr
    mgr._stop_event.set()


def test_on_scrape_completed_initializes_session(temp_manager):
    temp_manager.on_scrape_completed()
    assert temp_manager.is_active is True
    assert temp_manager.status == "active"
    assert temp_manager.update_count == 0
    assert temp_manager.tender_bids == []
    assert os.path.exists(temp_manager.live_excel_path)


def test_read_status_does_not_reset_timer(temp_manager):
    temp_manager.on_scrape_completed()
    t1 = temp_manager.last_updated_at
    time.sleep(0.05)
    status = temp_manager.get_status(touch=False)
    assert temp_manager.last_updated_at == t1
    assert status["is_active"] is True
    assert status["update_count"] == 0


def test_toggle_tender_adds_and_resets_timer(temp_manager):
    temp_manager.on_scrape_completed()
    t1 = temp_manager.last_updated_at
    time.sleep(0.02)
    res = temp_manager.toggle_tender("BID_001")
    assert res["action"] == "added"
    assert "BID_001" in temp_manager.tender_bids
    assert temp_manager.update_count == 1
    assert temp_manager.last_updated_at > t1


def test_toggle_tender_removes_if_present(temp_manager):
    temp_manager.on_scrape_completed()
    temp_manager.toggle_tender("BID_001")
    assert "BID_001" in temp_manager.tender_bids

    res = temp_manager.toggle_tender("BID_001")
    assert res["action"] == "removed"
    assert "BID_001" not in temp_manager.tender_bids
    assert temp_manager.update_count == 2


def test_toggle_forms_new_session_if_idle(temp_manager):
    assert temp_manager.is_active is False
    res = temp_manager.toggle_tender("BID_NEW")
    assert temp_manager.is_active is True
    assert "BID_NEW" in temp_manager.tender_bids
    assert temp_manager.update_count == 1


def test_zero_update_inactivity_discards_session(temp_manager):
    temp_manager.timeout_seconds = 0.3
    temp_manager.on_scrape_completed()
    assert temp_manager.is_active is True
    # Wait for watcher to trigger
    time.sleep(0.6)
    assert temp_manager.is_active is False
    assert temp_manager.status == "removed"
    assert not os.path.exists(temp_manager.live_excel_path)


def test_inactivity_with_updates_saves_sequentially(temp_manager):
    temp_manager.timeout_seconds = 0.3
    # First session with updates
    temp_manager.toggle_tender("BID_A")
    assert temp_manager.is_active is True
    time.sleep(0.6)
    assert temp_manager.is_active is False
    assert temp_manager.status == "saved"
    today = datetime.date.today().isoformat()
    expected_1 = f"{today}_1.xlsx"
    assert temp_manager.last_saved_filename == expected_1
    assert os.path.exists(os.path.join(temp_manager.daily_dir, expected_1))

    # Second session with updates -> forms 2nd file
    temp_manager.toggle_tender("BID_B")
    assert temp_manager.is_active is True
    time.sleep(0.6)
    assert temp_manager.is_active is False
    assert temp_manager.status == "saved"
    expected_2 = f"{today}_2.xlsx"
    assert temp_manager.last_saved_filename == expected_2
    assert os.path.exists(os.path.join(temp_manager.daily_dir, expected_2))


def test_build_curated_workbook_structure(tmp_path):
    output_file = str(tmp_path / "test_summary.xlsx")
    mock_tenders = [
        {
            "bid_no": "GEM/2026/B/100",
            "title": "Supply of Drones",
            "department": "Ministry of Defence",
            "quantity": "10",
            "end_date": "20-09-2026 05:00 PM",
            "status": "Shortlisted",
            "pdf_url": "https://gem.gov.in/bid/100",
            "analysis": {
                "recommendation": "Pursue",
                "priority_score": 85.0,
                "fit_score": 90.0,
                "score": 15.0,
                "est_value_inr": 2500000,
            }
        },
        {
            "bid_no": "GEM/2026/B/200",
            "title": "IT Equipment",
            "department": "Railways",
            "quantity": "50",
            "end_date": "25-09-2026 05:00 PM",
            "status": "Pending Review",
            "pdf_url": "https://gem.gov.in/bid/200",
            "analysis": {
                "recommendation": "Review",
                "priority_score": 60.0,
                "fit_score": 65.0,
                "score": 25.0,
                "est_value_inr": 1200000,
            }
        }
    ]

    path = build_curated_workbook(mock_tenders, output_file)
    assert os.path.exists(path)

    wb = openpyxl.load_workbook(path)
    sheet_names = wb.sheetnames
    assert "Overview" in sheet_names
    assert "Pursue" in sheet_names
    assert "Review" in sheet_names
    assert "Drop" in sheet_names
    assert "All_Tenders" in sheet_names

    # Check All_Tenders has 2 data rows
    ws_all = wb["All_Tenders"]
    # Row 1 is header, row 2 and 3 are data
    assert ws_all.max_row == 3
    assert ws_all.cell(row=1, column=6).value == "Bid Number"


def test_api_live_excel_routes():
    from app import app, live_excel_manager

    client = app.test_client()

    # 1. GET /api/live-excel
    res = client.get("/api/live-excel")
    assert res.status_code == 200
    data = res.get_json()
    assert "is_active" in data
    assert "seconds_remaining" in data
    assert "tender_bids" in data

    # 2. POST /api/live-excel/toggle
    res_toggle = client.post("/api/live-excel/toggle", json={"bid_no": "TEST/GEM/001"})
    assert res_toggle.status_code == 200
    tdata = res_toggle.get_json()
    assert tdata["action"] == "added"
    assert "TEST/GEM/001" in tdata["status"]["tender_bids"]

    # 3. POST /api/live-excel/add-batch
    res_batch = client.post("/api/live-excel/add-batch", json={"bid_nos": ["TEST/GEM/002", "TEST/GEM/003"]})
    assert res_batch.status_code == 200
    bdata = res_batch.get_json()
    assert "TEST/GEM/002" in bdata["tender_bids"]
    assert "TEST/GEM/003" in bdata["tender_bids"]

    # 4. GET /api/live-excel/download/live
    res_dl = client.get("/api/live-excel/download/live")
    assert res_dl.status_code == 200
    assert res_dl.headers["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # 5. POST /api/live-excel/close
    res_close = client.post("/api/live-excel/close", json={"save": True})
    assert res_close.status_code == 200
    cdata = res_close.get_json()
    assert "filename" in cdata
    assert cdata["status"]["is_active"] is False
