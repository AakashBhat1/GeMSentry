import os
import sys
import json
import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import openpyxl
from gemsentry.master_sheet import MasterSheetManager

@pytest.fixture
def temp_manager(tmp_path):
    mgr = MasterSheetManager()
    test_wb_path = str(tmp_path / "test_master.xlsx")
    
    # Create clean test workbook with the required sheets
    wb = openpyxl.Workbook()
    ws_master = wb.active
    ws_master.title = "MASTER"
    ws_study = wb.create_sheet("UNDER DETAILED STUDY")
    ws_part = wb.create_sheet("(TENDER DETAILS (PARTICIPATED)")
    
    # Add dummy header rows (row 4 for Master/Study, row 2 for Participated)
    for _ in range(3):
        ws_master.append([])
        ws_study.append([])
    ws_master.append(["SL. NO", "DOWNLOAD FROM", "WORK CATEGORY", "DOWNLOAD DATE", "MONTH", "ORGANISATION", "LOCATION/SITE", "TENDER ID"])
    ws_study.append(["SL. NO", "DOWNLOAD FROM", "WORK CATEGORY", "DOWNLOAD DATE", "MONTH", "ORGANISATION", "LOCATION/SITE", "TENDER ID"])
    
    ws_part.append([])
    ws_part.append(["SL. NO", "STATUS", "DOWNLOAD FROM", "WORK CATEGORY", "DOWNLOAD DATE", "MONTH", "ORGANISATION", "LOCATIOIN/SITE", "TENDER ID"])
    
    wb.save(test_wb_path)
    wb.close()

    mgr._get_active_master_paths = lambda: [test_wb_path]
    mgr.finalized_records = []
    mgr.highest_serial_no = 1016
    mgr._save_store = lambda: None
    mgr._sync_to_google_sheet = lambda payload: {"status": "ok", "mocked": True}
    return mgr


def test_serial_number_increment(temp_manager):
    sample_tender_1 = {
        "bid_no": "TEST/GEM/2026/001",
        "title": "Rugged Drone Detection System",
        "department": "Indian Air Force",
        "end_date": "15-06-2026 15:00:00",
        "analysis": {
            "buyer_org": "IAF 7 BRD",
            "est_value_inr": 4500000.0,
            "startup_exemption": "Yes (Full)",
            "mse_exemption": "Yes (Full)",
            "emd_status": "No EMD Required (OK)",
            "business_line": {"label": "DRONES"}
        }
    }

    sample_tender_2 = {
        "bid_no": "TEST/GEM/2026/002",
        "title": "AC-DC Linear Power Supply Units",
        "department": "Indian Navy",
        "end_date": "20-07-2026 11:00:00",
        "analysis": {
            "buyer_org": "Naval Dockyard",
            "est_value_inr": 1200000.0,
            "startup_exemption": "No",
            "mse_exemption": "Yes",
            "emd_amount": 25000.0,
            "business_line": {"label": "POWER SUPPLY"}
        }
    }

    # 1. Finalize tender 1
    res1 = temp_manager.finalize_tender(sample_tender_1, target_sheet="UNDER DETAILED STUDY")
    assert res1["status"] == "ok"
    assert res1["sl_no"] == 1017
    assert res1["record"]["bid_no"] == "TEST/GEM/2026/001"
    assert res1["record"]["organisation"] == "IAF 7 BRD"
    assert res1["record"]["experience_exemption"] == "YES"
    assert res1["record"]["work_category"] == "DRONES"

    # 2. Finalize tender 2
    res2 = temp_manager.finalize_tender(sample_tender_2, target_sheet="MASTER")
    assert res2["status"] == "ok"
    assert res2["sl_no"] == 1018
    assert res2["record"]["bid_no"] == "TEST/GEM/2026/002"
    assert res2["record"]["experience_exemption"] == "NO"
    assert res2["record"]["turnover_exemption"] == "YES"

    # 3. Check highest serial number
    assert temp_manager.get_highest_serial_number() == 1018


def test_move_to_participated(temp_manager):
    tender = {
        "bid_no": "TEST/GEM/2026/PART1",
        "title": "Specialized FPGA Boards",
        "department": "HAL Korwa",
        "analysis": {"est_value_inr": 850000.0}
    }
    temp_manager.finalize_tender(tender, target_sheet="UNDER DETAILED STUDY")

    # Move to participated with WON status
    res = temp_manager.move_to_participated(
        bid_no="TEST/GEM/2026/PART1",
        won_lost_result="WON L - 1",
        tender_value=850000.0,
        so_link="https://drive.google.com/file/d/test_so/view",
        final_remarks="Order placed successfully"
    )

    assert res["status"] == "ok"
    rec = temp_manager.get_record("TEST/GEM/2026/PART1")
    assert rec is not None
    assert rec["target_sheet"] == "(TENDER DETAILS (PARTICIPATED)"
    assert rec["won_lost_result"] == "WON L - 1"
    assert rec["so_status"] == "SO RECEIVED"
    assert rec["so_link"] == "https://drive.google.com/file/d/test_so/view"


def test_delete_tender(temp_manager):
    tender = {
        "bid_no": "TEST/GEM/2026/DEL",
        "title": "Accidental Tender",
        "department": "Test Dept"
    }
    temp_manager.finalize_tender(tender)
    assert temp_manager.is_tender_finalized("TEST/GEM/2026/DEL") is True

    # Delete
    del_res = temp_manager.delete_tender("TEST/GEM/2026/DEL")
    assert del_res["status"] == "ok"
    assert del_res["deleted_bid"] == "TEST/GEM/2026/DEL"
    assert temp_manager.is_tender_finalized("TEST/GEM/2026/DEL") is False


def test_api_finalized_endpoints(monkeypatch):
    from app import app, master_sheet_manager
    monkeypatch.setattr(master_sheet_manager, "_sync_to_local_excel", lambda *args, **kwargs: True)
    monkeypatch.setattr(master_sheet_manager, "_delete_from_local_excel", lambda *args, **kwargs: 1)
    monkeypatch.setattr(master_sheet_manager, "_sync_to_google_sheet", lambda *args, **kwargs: {"status": "ok", "mocked": True})

    client = app.test_client()

    # 1. GET /api/finalized
    res = client.get("/api/finalized")
    assert res.status_code == 200
    data = res.get_json()
    assert "highest_serial_no" in data
    assert "records" in data

    # 2. POST /api/finalized/finalize
    res_fin = client.post("/api/finalized/finalize", json={
        "bid_no": "API/TEST/GEM/001",
        "title": "API Test Tender",
        "target_sheet": "UNDER DETAILED STUDY"
    })
    assert res_fin.status_code == 200
    fin_data = res_fin.get_json()
    assert fin_data["status"] == "ok"
    assert fin_data["sl_no"] >= 1017

    # 3. POST /api/finalized/move-to-participated
    res_part = client.post("/api/finalized/move-to-participated", json={
        "bid_no": "API/TEST/GEM/001",
        "won_lost_result": "WON L - 1",
        "tender_value": 500000,
        "so_link": "https://drive.google.com/file/d/sample"
    })
    assert res_part.status_code == 200
    part_data = res_part.get_json()
    assert part_data["status"] == "ok"
    assert part_data["result"] == "WON L - 1"

    # 4. POST /api/finalized/delete
    res_del = client.post("/api/finalized/delete", json={
        "bid_no": "API/TEST/GEM/001"
    })
    assert res_del.status_code == 200
    del_data = res_del.get_json()
    assert del_data["status"] == "ok"

    # 5. GET /api/finalized/config
    res_cfg = client.get("/api/finalized/config")
    assert res_cfg.status_code == 200
    cfg_data = res_cfg.get_json()
    assert "spreadsheet_id" in cfg_data

    # 6. GET /api/finalized/script
    res_scr = client.get("/api/finalized/script")
    assert res_scr.status_code == 200
    scr_data = res_scr.get_json()
    assert "GeMSentry" in scr_data.get("script", "")

    # 7. POST /api/finalized/sync-all
    monkeypatch.setattr(master_sheet_manager, "config", {"apps_script_url": "https://script.google.com/macros/s/mock/exec"})
    res_sync = client.post("/api/finalized/sync-all")
    assert res_sync.status_code == 200
    sync_data = res_sync.get_json()
    assert sync_data["status"] == "ok"

