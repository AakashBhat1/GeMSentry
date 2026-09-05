"""
Master Sheet and Google Sheets Integration Engine for GeMSentry.

Manages:
1. Sequential serial numbering (SL. NO) starting from existing master numbering (>= 1016).
2. Row generation and formatting matching 'TENDER MASTER SHEET(ETSPL) 2025- 26.xlsx'.
3. Bidirectional tracking and persistence in data/finalized_tenders.json.
4. Writing to local Excel master workbooks (Downloads and repo copies).
5. Real-time syncing with Google Sheet and Google Drive via Google Apps Script Webhook.
6. Lifecycle transitions to '(TENDER DETAILS (PARTICIPATED)' with Won/Lost status.
7. Row deletion for correcting accidental additions.
"""

import os
import re
import json
import base64
import shutil
import logging
import datetime
import threading
from typing import Optional, List, Dict, Any, Tuple

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

import paths
from gemsentry.dateparse import parse_gem_date
from gemsentry.storage import find_existing_pdf_file

logger = logging.getLogger("gemsentry.master_sheet")

CONFIG_PATH = os.path.join(paths.CONFIG_DIR, "google_sync_config.json")
FINALIZED_STORE_PATH = os.path.join(paths.DATA_DIR, "finalized_tenders.json")
DEFAULT_LOCAL_MASTER_PATH = r"C:\Users\zewan\Downloads\TENDER MASTER SHEET(ETSPL) 2025- 26.xlsx"
WORKSPACE_MASTER_PATH = os.path.join(paths.ROOT, "TENDER MASTER SHEET(ETSPL) 2025- 26.xlsx")

# Baseline serial number if no previous records exist
BASELINE_SERIAL_NO = 1016

MASTER_COLUMNS = [
    "SL. NO", "DOWNLOAD FROM", "WORK CATEGORY", "DOWNLOAD DATE", "MONTH",
    "ORGANISATION", "LOCATION/SITE", "TENDER ID", "REFERENCE NO.", "DESCRIPTION",
    "BID SUBMISSION (END DATE)", "BID SUBMISSION (END TIME)", "EXPERIENCE EXEMPTION\nYES/ NO",
    "TURNOVER EXEMPTION\nYES/ NO", "EMD/ TENDER FEES", "OEM AUTHORIZATION", "RFP LINK",
    "APPROVAL", "REMARKS"
]

PARTICIPATED_COLUMNS = [
    "SL. NO", "STATUS", "DOWNLOAD FROM", "WORK CATEGORY", "DOWNLOAD DATE", "MONTH",
    "ORGANISATION", "LOCATIOIN/SITE", "TENDER ID", "REFERENCENO.", "DESCRIPTION",
    "BID SUBMISSION (END DATE)", "BID SUBMISSION (END TIME)", "BID OPENING DATE",
    "SUBMISSION STATUS", "SUBMITTED BY", "REMARKS", "JOB ALIGNED TO", "ETSPL CTC",
    "TENDER VALUE", "EMD/ TRANSACTION/ DOCUMENT", "TECHNICAL STATUS", "FINANCIAL STATUS",
    "RESULT\nWON/LOST", "SO/ DO  STATUS", "SO LINK", "REMARKS"
]


class MasterSheetManager:
    """Coordinates finalized tenders between GeMSentry, Excel master sheets, and Google Sheets."""

    def __init__(self):
        self.lock = threading.RLock()
        self.config = self._load_config()
        self.finalized_records: List[Dict[str, Any]] = self._load_store()
        self._ensure_serial_baseline()

    def _load_config(self) -> Dict[str, Any]:
        default_cfg = {
            "spreadsheet_id": "1WbeJJ8goLPGLryyJfcJNbiXtIxXjC9Z0g8viueh5oOk",
            "spreadsheet_url": "https://docs.google.com/spreadsheets/d/1WbeJJ8goLPGLryyJfcJNbiXtIxXjC9Z0g8viueh5oOk/edit?usp=sharing",
            "apps_script_url": "",
            "google_drive_mount_path": "",
            "local_master_excel_path": DEFAULT_LOCAL_MASTER_PATH,
            "sync_to_local_excel": True,
            "sync_to_google_sheet": True,
            "default_sheet": "UNDER DETAILED STUDY"
        }
        if os.path.exists(CONFIG_PATH):
            try:
                with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                    user_cfg = json.load(f)
                    default_cfg.update(user_cfg)
            except Exception as e:
                logger.warning("Could not read google_sync_config.json: %s", e)
        return default_cfg

    def load_config(self) -> Dict[str, Any]:
        """Reloads and returns the latest config from disk."""
        with self.lock:
            self.config = self._load_config()
            return self.config

    def save_config(self, new_config: Dict[str, Any]) -> Dict[str, Any]:
        with self.lock:
            self.config.update(new_config)
            os.makedirs(os.path.dirname(CONFIG_PATH), exist_ok=True)
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(self.config, f, indent=2)
            logger.info("Updated Google sync configuration.")
            return self.config

    def _load_store(self) -> List[Dict[str, Any]]:
        if os.path.exists(FINALIZED_STORE_PATH):
            try:
                with open(FINALIZED_STORE_PATH, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        return data
            except Exception as e:
                logger.warning("Failed loading finalized_tenders.json: %s", e)
        return []

    def _save_store(self):
        os.makedirs(os.path.dirname(FINALIZED_STORE_PATH), exist_ok=True)
        try:
            with open(FINALIZED_STORE_PATH, "w", encoding="utf-8") as f:
                json.dump(self.finalized_records, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logger.error("Failed persisting finalized_tenders.json: %s", e)

    def _get_active_master_paths(self) -> List[str]:
        paths_to_update = []
        configured_path = self.config.get("local_master_excel_path") or DEFAULT_LOCAL_MASTER_PATH
        if os.path.exists(configured_path):
            paths_to_update.append(configured_path)
        if os.path.exists(WORKSPACE_MASTER_PATH) and WORKSPACE_MASTER_PATH not in paths_to_update:
            paths_to_update.append(WORKSPACE_MASTER_PATH)
        return paths_to_update

    def _ensure_serial_baseline(self):
        """Scans local excel files to determine the current highest serial number from actual tenders."""
        max_sl = BASELINE_SERIAL_NO
        for rec in self.finalized_records:
            sl = rec.get("sl_no")
            if isinstance(sl, (int, float)) and sl > max_sl:
                max_sl = int(sl)

        for path in self._get_active_master_paths():
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                for sname in ["MASTER", "UNDER DETAILED STUDY", "(TENDER DETAILS (PARTICIPATED)"]:
                    if sname in wb.sheetnames:
                        sheet = wb[sname]
                        for row in sheet.iter_rows(values_only=True):
                            # Only consider rows that actually have a Tender ID (col index 7 or 8)
                            if len(row) > 7 and row[7]:
                                val_id = str(row[7]).strip()
                                if val_id and val_id.upper() != "TENDER ID":
                                    val_sl = row[0]
                                    if val_sl is not None:
                                        try:
                                            val = int(float(str(val_sl).strip()))
                                            if BASELINE_SERIAL_NO <= val < 50000 and val > max_sl:
                                                max_sl = val
                                        except (ValueError, TypeError):
                                            pass
                wb.close()
            except Exception as e:
                logger.debug("Error checking baseline serial from %s: %s", path, e)

        self.highest_serial_no = max_sl

    def get_highest_serial_number(self) -> int:
        with self.lock:
            max_sl = getattr(self, "highest_serial_no", BASELINE_SERIAL_NO)
            for rec in self.finalized_records:
                sl = rec.get("sl_no")
                if isinstance(sl, (int, float)) and sl > max_sl:
                    max_sl = int(sl)
            return max_sl

    def is_tender_finalized(self, bid_no: str) -> bool:
        norm_bid = str(bid_no).strip().lower()
        return any(str(r.get("bid_no")).strip().lower() == norm_bid for r in self.finalized_records)

    def get_record(self, bid_no: str) -> Optional[Dict[str, Any]]:
        norm_bid = str(bid_no).strip().lower()
        for r in self.finalized_records:
            if str(r.get("bid_no")).strip().lower() == norm_bid:
                return r
        return None

    def _format_date_parts(self, date_str: Optional[str]) -> Tuple[str, str]:
        if not date_str:
            return "N/A", "15:00"
        dt = parse_gem_date(date_str)
        if dt:
            return dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M")
        return str(date_str)[:10], "15:00"

    def _handle_google_drive(self, tender: Dict[str, Any], bid_no: str) -> str:
        """Handles PDF copy to mounted Google Drive or cloud upload."""
        # 1. Reuse existing Google Drive link if already present
        existing_drive = tender.get("rfp_link") or tender.get("drive_link")
        if existing_drive and "drive.google.com" in str(existing_drive):
            return str(existing_drive)

        local_pdf = tender.get("local_pdf_path")
        abs_pdf = ""
        if local_pdf:
            abs_pdf = local_pdf if os.path.isabs(local_pdf) else os.path.join(paths.ROOT, local_pdf)
            if not os.path.exists(abs_pdf):
                abs_pdf = ""

        # Auto-search disk for PDF if not specified
        if not abs_pdf:
            sanitized = re.sub(r'[\\/*?:"<>|]', "_", bid_no)
            found = find_existing_pdf_file(sanitized)
            if found:
                abs_pdf = os.path.join(paths.ROOT, found) if not os.path.isabs(found) else found

        # 2. Check local mounted drive folder (e.g. G:\My Drive\...)
        mount_path = (self.config.get("google_drive_mount_path") or "").strip()
        if mount_path and os.path.exists(mount_path) and os.path.isdir(mount_path) and abs_pdf and os.path.exists(abs_pdf):
            try:
                safe_name = re.sub(r'[\\/*?:"<>|]', "_", bid_no) + ".pdf"
                dest = os.path.join(mount_path, safe_name)
                shutil.copy2(abs_pdf, dest)
                logger.info("Copied tender PDF to mounted Google Drive: %s", dest)
                return dest
            except Exception as e:
                logger.warning("Failed copying PDF to mounted Google Drive: %s", e)

        # 3. Check Apps Script direct drive uploader
        apps_script_url = (self.config.get("apps_script_url") or "").strip()
        if apps_script_url and abs_pdf and os.path.exists(abs_pdf):
            try:
                with open(abs_pdf, "rb") as f:
                    pdf_b64 = base64.b64encode(f.read()).decode("utf-8")
                resp = requests.post(
                    apps_script_url,
                    json={
                        "action": "upload_pdf_to_drive",
                        "filename": f"{re.sub(r'[\\\\/*?:\"<>|]', '_', bid_no)}.pdf",
                        "base64_data": pdf_b64
                    },
                    timeout=20
                )
                data = resp.json()
                if data.get("drive_link"):
                    logger.info("Uploaded tender PDF to Google Drive via Apps Script: %s", data["drive_link"])
                    return data["drive_link"]
            except Exception as e:
                logger.warning("Failed uploading PDF to Google Drive via Apps Script: %s", e)

        # 4. Fallback: GeM PDF link or local relative path
        return tender.get("pdf_url") or local_pdf or ""

    def _sync_to_local_excel(self, record: Dict[str, Any], target_sheet: str) -> bool:
        """Writes row to local Excel files, respecting empty rows, header offsets, and styling neatly."""
        success = True
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for path in self._get_active_master_paths():
            try:
                wb = openpyxl.load_workbook(path)
                if target_sheet not in wb.sheetnames:
                    logger.warning("Sheet %s not found in %s", target_sheet, path)
                    wb.close()
                    continue
                ws = wb[target_sheet]

                header_row = 2 if "PARTICIPATED" in target_sheet else 4

                # 1. Search if tender already exists in sheet to update in-place
                target_row = None
                norm_bid = str(record.get("bid_no") or "").strip().lower()
                target_sl = record.get("sl_no")
                col_bid = 9 if "PARTICIPATED" in target_sheet else 8
                col_ref = 10 if "PARTICIPATED" in target_sheet else 9

                for r in range(header_row + 1, ws.max_row + 1):
                    cell_sl = ws.cell(row=r, column=1).value
                    cell_id = ws.cell(row=r, column=col_bid).value
                    cell_ref = ws.cell(row=r, column=col_ref).value

                    id_match = norm_bid and (
                        (cell_id and str(cell_id).strip().lower() == norm_bid) or
                        (cell_ref and str(cell_ref).strip().lower() == norm_bid)
                    )
                    sl_match = target_sl and cell_sl and str(cell_sl).strip() == str(target_sl).strip()
                    if id_match or sl_match:
                        target_row = r
                        break

                # 2. If not existing, find first genuinely empty row
                if not target_row:
                    for r in range(header_row + 1, ws.max_row + 2):
                        val_id = ws.cell(row=r, column=col_bid).value
                        val_desc = ws.cell(row=r, column=col_bid + 2).value
                        if val_id is None and val_desc is None:
                            target_row = r
                            break

                if not target_row:
                    target_row = ws.max_row + 1

                # Set generous 36pt row height for neat visibility
                ws.row_dimensions[target_row].height = 36.0

                # Populate row cells
                if "PARTICIPATED" in target_sheet:
                    row_data = [
                        record.get("sl_no"),
                        record.get("tender_type", "RFP"),
                        record.get("download_from", "GEM"),
                        record.get("work_category", "SUPPLY"),
                        record.get("download_date"),
                        record.get("month"),
                        record.get("organisation"),
                        record.get("location"),
                        record.get("bid_no"),
                        record.get("bid_no"),
                        record.get("title"),
                        record.get("end_date"),
                        record.get("end_time"),
                        record.get("drive_link") or record.get("rfp_link") or "",
                        record.get("submission_status", "SUBMITTED"),
                        record.get("submitted_by", "SUBMITTED BY ETSPL"),
                        record.get("remarks", ""),
                        record.get("job_aligned_to", ""),
                        record.get("etspl_ctc", ""),
                        record.get("tender_value", "N/A"),
                        record.get("emd_doc", "EXEMPTED"),
                        record.get("technical_status", "QUALIFIED"),
                        record.get("financial_status", "QUALIFIED"),
                        record.get("won_lost_result", "WON L - 1"),
                        record.get("so_status", "SO RECEIVED"),
                        record.get("so_link", ""),
                        record.get("final_remarks", "")
                    ]
                else:
                    row_data = [
                        record.get("sl_no"),
                        record.get("download_from", "GEM"),
                        record.get("work_category", "SUPPLY"),
                        record.get("download_date"),
                        record.get("month"),
                        record.get("organisation"),
                        record.get("location"),
                        record.get("bid_no"),
                        record.get("bid_no"),
                        record.get("title"),
                        record.get("end_date"),
                        record.get("end_time"),
                        record.get("experience_exemption", "YES"),
                        record.get("turnover_exemption", "YES"),
                        record.get("emd", 0.0),
                        record.get("oem_authorization", "YES"),
                        record.get("rfp_link", ""),
                        record.get("approval", "TO BE SUBMIT"),
                        record.get("remarks", "")
                    ]

                is_part = "PARTICIPATED" in target_sheet
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=target_row, column=col_idx, value=val)
                    cell.font = Font(name="Arial", size=11)
                    cell.border = thin_border

                    # Alignments & formatting
                    if not is_part:
                        if col_idx in (1, 2, 3, 4, 5, 8, 9, 11, 12, 13, 14, 16, 17, 18):
                            cell.alignment = Alignment(vertical="center", horizontal="center")
                        elif col_idx in (6, 10, 19):
                            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                        elif col_idx == 15:
                            cell.alignment = Alignment(vertical="center", horizontal="right")
                            cell.number_format = '#,##0'

                        if col_idx in (1, 8):
                            cell.font = Font(name="Arial", size=11, bold=True)
                        elif col_idx == 17 and val and str(val).startswith("http"):
                            cell.hyperlink = str(val)
                            cell.value = "Google Drive RFP ↗" if "drive.google.com" in str(val) else "Open RFP ↗"
                            cell.font = Font(name="Arial", size=11, color="1D4ED8", underline="single", bold=True)
                        elif col_idx == 18:
                            cell.font = Font(name="Arial", size=11, color="047857", bold=True)
                    else:
                        if col_idx in (1, 2, 3, 4, 5, 9, 10, 12, 13, 14, 15, 24, 25, 26):
                            cell.alignment = Alignment(vertical="center", horizontal="center")
                        elif col_idx in (7, 8, 11, 17, 27):
                            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                        elif col_idx in (20, 21):
                            cell.alignment = Alignment(vertical="center", horizontal="right")
                            cell.number_format = '#,##0'

                        if col_idx in (1, 9):
                            cell.font = Font(name="Arial", size=11, bold=True)
                        elif col_idx == 14 and val and str(val).startswith("http"):
                            cell.hyperlink = str(val)
                            cell.value = "Google Drive RFP ↗" if "drive.google.com" in str(val) else "Open RFP ↗"
                            cell.font = Font(name="Arial", size=11, color="1D4ED8", underline="single", bold=True)
                        elif col_idx == 24:
                            is_won = "WON" in str(val).upper()
                            cell.font = Font(name="Arial", size=11, color="047857" if is_won else "DC2626", bold=True)
                        elif col_idx == 26 and val and str(val).startswith("http"):
                            cell.hyperlink = str(val)
                            cell.value = "Open SO Doc ↗"
                            cell.font = Font(name="Arial", size=11, color="1D4ED8", underline="single", bold=True)

                wb.save(path)
                wb.close()
                logger.info("Successfully wrote row %d to sheet %s in %s", target_row, target_sheet, path)
            except PermissionError:
                logger.warning("Could not write to %s because it is open in another program.", path)
                success = False
            except Exception as e:
                logger.error("Error writing to local Excel workbook %s: %s", path, e)
                success = False
        return success

    def _delete_from_local_excel(self, bid_no: str, sl_no: Optional[int] = None) -> int:
        deleted_count = 0
        norm_bid = str(bid_no).strip().lower()
        for path in self._get_active_master_paths():
            try:
                wb = openpyxl.load_workbook(path)
                for sname in ["UNDER DETAILED STUDY", "MASTER", "(TENDER DETAILS (PARTICIPATED)"]:
                    if sname not in wb.sheetnames:
                        continue
                    ws = wb[sname]
                    header_row = 2 if "PARTICIPATED" in sname else 4
                    # Iterate backwards from bottom to header
                    for r in range(ws.max_row, header_row, -1):
                        cell_id = ws.cell(row=r, column=8).value
                        cell_ref = ws.cell(row=r, column=9).value
                        cell_sl = ws.cell(row=r, column=1).value

                        id_match = norm_bid and (
                            (cell_id and str(cell_id).strip().lower() == norm_bid) or
                            (cell_ref and str(cell_ref).strip().lower() == norm_bid)
                        )
                        sl_match = sl_no and cell_sl and int(float(str(cell_sl))) == sl_no

                        if id_match or sl_match:
                            ws.delete_rows(r, 1)
                            deleted_count += 1
                            logger.info("Deleted row %d from %s in %s", r, sname, path)

                wb.save(path)
                wb.close()
            except Exception as e:
                logger.error("Error deleting tender from local Excel %s: %s", path, e)
        return deleted_count

    def _sync_to_google_sheet(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """Sends action payload to Google Apps Script Webhook."""
        self.load_config()
        apps_script_url = (self.config.get("apps_script_url") or "").strip()
        if not apps_script_url:
            return {"status": "skipped", "message": "No Google Apps Script Webhook URL configured."}
        try:
            resp = requests.post(apps_script_url, json=payload, timeout=12)
            data = resp.json()
            logger.info("Google Sheet sync response: %s", data)
            return data
        except Exception as e:
            logger.warning("Failed communicating with Google Sheet Webhook: %s", e)
            return {"status": "error", "message": str(e)}

    def _build_gsheet_payload(
        self,
        record: Dict[str, Any],
        target_sheet: str = "MASTER",
        secondary_sheet: Optional[str] = None
    ) -> Dict[str, Any]:
        """Builds a comprehensive payload for Google Apps Script with both flat fields and nested tender."""
        rfp = record.get("rfp_link") or record.get("drive_link") or record.get("pdf_url") or ""
        return {
            "action": "append_tender",
            "target_sheet": target_sheet,
            "secondary_sheet": secondary_sheet,
            "sl_no": record.get("sl_no"),
            "download_from": record.get("download_from", "GEM"),
            "work_category": record.get("work_category", "SUPPLY"),
            "download_date": record.get("download_date"),
            "month": record.get("month"),
            "organisation": record.get("organisation"),
            "location": record.get("location"),
            "tender_id": record.get("bid_no"),
            "reference_no": record.get("bid_no"),
            "description": record.get("title"),
            "end_date": record.get("end_date"),
            "end_time": record.get("end_time"),
            "experience_exemption": record.get("experience_exemption", "YES"),
            "turnover_exemption": record.get("turnover_exemption", "YES"),
            "emd": record.get("emd", 0.0),
            "oem_authorization": record.get("oem_authorization", "YES"),
            "rfp_link": rfp,
            "approval": record.get("approval", "TO BE SUBMIT"),
            "remarks": record.get("remarks", ""),
            "tender": record
        }

    def finalize_tender(
        self,
        tender: Dict[str, Any],
        target_sheet: str = "UNDER DETAILED STUDY",
        custom_fields: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Finalizes a tender, assigns sequential SL. NO, updates Excel & Google Sheet."""
        with self.lock:
            bid_no = tender.get("bid_no") or "UNKNOWN"
            custom_fields = custom_fields or {}

            # Check if already finalized
            existing = self.get_record(bid_no)
            if existing:
                sl_no = existing.get("sl_no")
            else:
                sl_no = self.get_highest_serial_number() + 1
                self.highest_serial_no = sl_no

            analysis = tender.get("analysis") or {}
            now = datetime.datetime.now()
            end_date, end_time = self._format_date_parts(tender.get("end_date"))

            # Handle Google Drive
            rfp_link = self._handle_google_drive(tender, bid_no)

            # Exemptions & EMD
            exp_exempt = "YES" if "yes" in str(analysis.get("startup_exemption", "")).lower() else "NO"
            trn_exempt = "YES" if "yes" in str(analysis.get("mse_exemption", "")).lower() else "NO"
            emd_val = analysis.get("emd_amount") or 0.0
            if "no emd" in str(analysis.get("emd_status", "")).lower():
                emd_val = 0.0

            # Work Category
            work_cat = (
                custom_fields.get("work_category") or
                (analysis.get("business_line") or {}).get("label") or
                tender.get("nlp_category") or
                tender.get("item_category") or
                "SUPPLY"
            ).upper()

            record = {
                "sl_no": sl_no,
                "bid_no": bid_no,
                "download_from": (tender.get("source_name") or tender.get("source_id") or "GEM").upper(),
                "work_category": work_cat,
                "download_date": now.strftime("%Y-%m-%d"),
                "month": now.strftime("%B").upper(),
                "organisation": analysis.get("buyer_org") or tender.get("department") or "N/A",
                "location": analysis.get("consignee_state") or "N/A",
                "title": tender.get("title") or analysis.get("primary_item") or "N/A",
                "end_date": end_date,
                "end_time": end_time,
                "experience_exemption": exp_exempt,
                "turnover_exemption": trn_exempt,
                "emd": emd_val,
                "oem_authorization": custom_fields.get("oem_authorization") or "YES",
                "rfp_link": rfp_link,
                "approval": custom_fields.get("approval") or "TO BE SUBMIT",
                "remarks": custom_fields.get("remarks") or (f"Pre-Bid: {analysis.get('pre_bid_date')}" if analysis.get("pre_bid_date") else ""),
                "target_sheet": target_sheet,
                "finalized_at": now.isoformat(),
                "est_value_inr": analysis.get("est_value_inr") or tender.get("est_value_inr")
            }

            # Update records
            if existing:
                existing.update(record)
            else:
                self.finalized_records.append(record)
            self._save_store()

            # Local Excel Sync - Compulsorily append to MASTER, and also to target_sheet if different
            excel_synced = self._sync_to_local_excel(record, "MASTER")
            if target_sheet and target_sheet != "MASTER":
                self._sync_to_local_excel(record, target_sheet)

            # Google Sheet Sync - Compulsorily append to MASTER
            gsheet_payload = self._build_gsheet_payload(
                record=record,
                target_sheet="MASTER",
                secondary_sheet=target_sheet if target_sheet != "MASTER" else None
            )
            gsheet_res = self._sync_to_google_sheet(gsheet_payload)

            return {
                "status": "ok",
                "sl_no": sl_no,
                "bid_no": bid_no,
                "record": record,
                "local_excel_synced": excel_synced,
                "google_sheet_synced": gsheet_res.get("status") == "ok",
                "google_response": gsheet_res
            }

    def delete_tender(self, bid_no_or_sl_no: Any) -> Dict[str, Any]:
        """Deletes a finalized tender from JSON store, local Excel, and Google Sheet."""
        with self.lock:
            target_bid = None
            target_sl = None

            # Match in store
            matched_idx = -1
            for idx, r in enumerate(self.finalized_records):
                if str(r.get("bid_no")).strip().lower() == str(bid_no_or_sl_no).strip().lower():
                    matched_idx = idx
                    target_bid = r.get("bid_no")
                    target_sl = r.get("sl_no")
                    break
                if str(r.get("sl_no")) == str(bid_no_or_sl_no):
                    matched_idx = idx
                    target_bid = r.get("bid_no")
                    target_sl = r.get("sl_no")
                    break

            if matched_idx >= 0:
                self.finalized_records.pop(matched_idx)
                self._save_store()
            else:
                target_bid = str(bid_no_or_sl_no)

            # Delete from Excel
            excel_del = self._delete_from_local_excel(target_bid, target_sl)

            # Delete from Google Sheet
            gsheet_res = self._sync_to_google_sheet({
                "action": "delete_tender",
                "bid_no": target_bid,
                "sl_no": target_sl
            })

            return {
                "status": "ok",
                "deleted_bid": target_bid,
                "deleted_sl": target_sl,
                "excel_rows_deleted": excel_del,
                "google_response": gsheet_res
            }

    def move_to_participated(
        self,
        bid_no: str,
        won_lost_result: str = "WON L - 1",
        tender_value: Optional[Any] = None,
        so_link: Optional[str] = None,
        submission_status: str = "SUBMITTED",
        final_remarks: Optional[str] = None
    ) -> Dict[str, Any]:
        """Transitions a finalized tender to '(TENDER DETAILS (PARTICIPATED)'."""
        with self.lock:
            record = self.get_record(bid_no)
            if not record:
                return {"status": "error", "message": f"Tender {bid_no} is not finalized yet."}

            record["tender_type"] = "RFP"
            record["won_lost_result"] = won_lost_result
            record["submission_status"] = submission_status
            record["submitted_by"] = "SUBMITTED BY ETSPL"
            record["tender_value"] = tender_value or record.get("est_value_inr") or "N/A"
            record["so_status"] = "SO RECEIVED" if "won" in won_lost_result.lower() else "SUBMITTED"
            record["so_link"] = so_link or ""
            record["final_remarks"] = final_remarks or ""
            record["target_sheet"] = "(TENDER DETAILS (PARTICIPATED)"
            self._save_store()

            # Append to Participated sheet in Excel
            excel_synced = self._sync_to_local_excel(record, "(TENDER DETAILS (PARTICIPATED)")

            # Send to Google Sheet
            gsheet_res = self._sync_to_google_sheet({
                "action": "move_to_participated",
                "tender_id": bid_no,
                "sl_no": record.get("sl_no"),
                "won_lost_result": won_lost_result,
                "tender_value": record.get("tender_value"),
                "so_link": record.get("so_link"),
                "tender": record
            })

            return {
                "status": "ok",
                "bid_no": bid_no,
                "sl_no": record.get("sl_no"),
                "result": won_lost_result,
                "local_excel_synced": excel_synced,
                "google_response": gsheet_res
            }

    def sync_all_to_google_sheet(self) -> Dict[str, Any]:
        """Pushes all finalized tenders from local store into Google Sheet."""
        with self.lock:
            apps_script_url = (self.config.get("apps_script_url") or "").strip()
            if not apps_script_url:
                return {
                    "status": "error",
                    "message": "No Google Apps Script Webhook URL configured in Settings. Please deploy the Apps Script as a Web App and paste its URL."
                }

            if not self.finalized_records:
                return {
                    "status": "ok",
                    "message": "No finalized tenders to sync.",
                    "synced_count": 0
                }

            synced = []
            errors = []
            for record in self.finalized_records:
                target_sheet = record.get("target_sheet") or "UNDER DETAILED STUDY"
                # Ensure local Excel has it in MASTER
                self._sync_to_local_excel(record, "MASTER")
                if target_sheet and target_sheet != "MASTER":
                    self._sync_to_local_excel(record, target_sheet)

                payload = self._build_gsheet_payload(
                    record=record,
                    target_sheet="MASTER",
                    secondary_sheet=target_sheet if target_sheet != "MASTER" else None
                )
                res = self._sync_to_google_sheet(payload)
                if res.get("status") == "ok":
                    synced.append(record.get("sl_no"))
                else:
                    errors.append({
                        "sl_no": record.get("sl_no"),
                        "bid_no": record.get("bid_no"),
                        "error": res.get("message") or res.get("error") or str(res)
                    })

            return {
                "status": "ok",
                "synced_count": len(synced),
                "synced_sl_nos": synced,
                "errors": errors,
                "message": f"Successfully synced {len(synced)} of {len(self.finalized_records)} tender(s) to Google Sheet."
            }

    def get_all_finalized(self) -> Dict[str, Any]:
        with self.lock:
            sorted_records = sorted(
                self.finalized_records,
                key=lambda r: int(r.get("sl_no") or 0),
                reverse=True
            )
            sheets_count = {}
            for r in self.finalized_records:
                s = r.get("target_sheet", "UNDER DETAILED STUDY")
                sheets_count[s] = sheets_count.get(s, 0) + 1

            return {
                "total_count": len(self.finalized_records),
                "highest_serial_no": self.get_highest_serial_number(),
                "records": sorted_records,
                "counts_by_sheet": sheets_count,
                "spreadsheet_url": self.config.get("spreadsheet_url"),
                "spreadsheet_id": self.config.get("spreadsheet_id"),
                "has_webhook": bool(self.config.get("apps_script_url")),
                "has_gdrive_mount": bool(self.config.get("google_drive_mount_path"))
            }


master_sheet_manager = MasterSheetManager()

