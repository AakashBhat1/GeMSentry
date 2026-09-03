"""
Live Excel Session Manager.

Implements the 10-minute inactivity lifecycle for interactive tender curation:
1. An Excel session is stored after each new scrape.
2. If stored for > 10 minutes with NO updates, it is automatically removed/discarded.
3. Updates (adding, removing, or modifying tenders via card clickers) reset the 10-minute timer.
4. Reads (checking status, viewing, or downloading) do NOT count as updates and do NOT reset the timer.
5. When 10 minutes elapse without further updates (and tenders were added), the session closes
   and is saved permanently as `<YYYY-MM-DD>_<N>.xlsx` (e.g. `2026-09-04_1.xlsx`).
6. If the 10 minutes have expired and a user selects a tender card to add to Excel, a new
   Excel session is formed following the exact same 10-minute lifecycle.
"""

import os
import re
import glob
import json
import time
import shutil
import logging
import datetime
import threading
import urllib.parse
from typing import Optional, List, Dict, Any, Set

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import paths

logger = logging.getLogger("gemsentry.live_excel")

# Default 10 minutes (600 seconds) inactivity timeout
DEFAULT_TIMEOUT_SECONDS = 600

# Styling constants matching tools/export_summary.py
HEADERS = [
    ("Priority", 9), ("Fit", 6), ("Risk", 6), ("Recommendation", 14),
    ("Status", 14), ("Bid Number", 20), ("Title", 60), ("Business Line", 18),
    ("Buyer / Department", 40), ("Est. Value (INR)", 15), ("Days Left", 9),
    ("End Date", 18), ("EMD", 22), ("Startup Exemption", 16),
    ("MSE Exemption", 16), ("Confidence", 10), ("Keyword", 22),
    ("GeM Portal", 11), ("Local PDF", 10),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RED_FILL = PatternFill("solid", fgColor="F8CBAD")
AMBER_FILL = PatternFill("solid", fgColor="FFE699")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
LINK_FONT = Font(color="0563C1", underline="single")


def _days_left(tender: Dict[str, Any]) -> Optional[float]:
    end_str = tender.get("end_date")
    if not end_str:
        return None
    try:
        from gemsentry.dateparse import parse_gem_date
        end = parse_gem_date(end_str)
        if not end:
            return None
        return round((end - datetime.datetime.now()).total_seconds() / 86400.0, 1)
    except Exception:
        return None


def _file_uri(rel_path: Optional[str]) -> Optional[str]:
    if not rel_path:
        return None
    abs_path = rel_path if os.path.isabs(rel_path) else os.path.join(paths.ROOT, rel_path)
    if not os.path.exists(abs_path):
        return None
    return "file:///" + urllib.parse.quote(abs_path.replace("\\", "/"), safe="/:")


def _sort_key(tender: Dict[str, Any]) -> float:
    pr = (tender.get("analysis") or {}).get("priority_score")
    return float(pr) if pr is not None else -1.0


def _write_sheet(wb, name: str, tenders: List[Dict[str, Any]], tab_color: Optional[str] = None):
    ws = wb.create_sheet(name)
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    for col, (title, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    for row, t in enumerate(sorted(tenders, key=_sort_key, reverse=True), 2):
        a = t.get("analysis") or {}
        dl = _days_left(t)
        rec = a.get("recommendation")
        bl = a.get("business_line") or {}
        values = [
            a.get("priority_score"), a.get("fit_score"), a.get("score"),
            rec or "—", t.get("status"), t.get("bid_no"),
            (t.get("title") or "")[:200], bl.get("label") or "—",
            (t.get("department") or a.get("buyer_org") or "")[:120],
            a.get("est_value_inr"), dl, t.get("end_date") or "—",
            a.get("emd_status") or "—", a.get("startup_exemption") or "—",
            a.get("mse_exemption") or "—", a.get("confidence"),
            t.get("keyword") or "", None, None,
        ]
        for col, v in enumerate(values, 1):
            ws.cell(row=row, column=col, value=v)

        if rec == "Pursue":
            ws.cell(row=row, column=4).fill = GREEN_FILL
        if dl is not None and dl <= 10:
            ws.cell(row=row, column=11).fill = RED_FILL if dl <= 5 else AMBER_FILL
        if a.get("est_value_inr") is not None:
            ws.cell(row=row, column=10).number_format = "#,##0"

        gem = ws.cell(row=row, column=18)
        if t.get("pdf_url"):
            gem.value = "Open Bid"
            gem.hyperlink = t["pdf_url"]
            gem.font = LINK_FONT
        else:
            gem.value = "—"
        pdf = ws.cell(row=row, column=19)
        uri = _file_uri(t.get("local_pdf_path"))
        if uri:
            pdf.value = "Open PDF"
            pdf.hyperlink = uri
            pdf.font = LINK_FONT
        else:
            pdf.value = "—"
    return ws


def _write_overview(wb, tenders: List[Dict[str, Any]], fingerprint: Optional[str] = None):
    ws = wb.create_sheet("Overview", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    title = ws.cell(row=1, column=1, value="GeM Tender Curated Summary")
    title.font = Font(size=16, bold=True)
    rows = [
        ("Generated", datetime.datetime.now().strftime("%d %b %Y, %I:%M %p")),
        ("Total selected tenders", len(tenders)),
        ("Scoring fingerprint", fingerprint or "—"),
        ("", ""),
    ]
    recs, statuses = {}, {}
    for t in tenders:
        r = (t.get("analysis") or {}).get("recommendation") or "Unanalyzed"
        recs[r] = recs.get(r, 0) + 1
        s = t.get("status") or "—"
        statuses[s] = statuses.get(s, 0) + 1
    rows += [("By recommendation", "")] + sorted(recs.items())
    rows += [("", ""), ("By status", "")] + sorted(statuses.items())
    for i, (k, v) in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=(v == ""))
        ws.cell(row=i, column=2, value=v)
    return ws


def build_curated_workbook(tenders: List[Dict[str, Any]], output_path: str):
    """Generate a full summary workbook for the selected tenders."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    try:
        from gemsentry.config_store import load_scoring_config
        from gemsentry.profile import load_company_profile
        from gemsentry.scoring.verdict import scoring_fingerprint
        cfg = load_scoring_config()
        profile = load_company_profile()
        fp = scoring_fingerprint(cfg, profile)
    except Exception:
        fp = None

    by_rec = {"Pursue": [], "Review": [], "Drop": []}
    for t in tenders:
        rec = (t.get("analysis") or {}).get("recommendation")
        if rec in by_rec:
            by_rec[rec].append(t)
        else:
            by_rec.setdefault(rec or "Other", []).append(t)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    _write_overview(wb, tenders, fp)
    _write_sheet(wb, "Pursue", by_rec.get("Pursue", []), tab_color="2E7D32")
    _write_sheet(wb, "Review", by_rec.get("Review", []), tab_color="F9A825")
    _write_sheet(wb, "Drop", by_rec.get("Drop", []), tab_color="C62828")
    _write_sheet(wb, "All_Tenders", tenders, tab_color="1565C0")

    wb.save(output_path)
    return output_path


class LiveExcelManager:
    """
    Manages active live Excel sessions, inactivity timeouts, and persistent daily exports.
    """

    def __init__(self, timeout_seconds: int = DEFAULT_TIMEOUT_SECONDS):
        self.timeout_seconds = timeout_seconds
        self.lock = threading.RLock()

        self.reports_dir = os.path.join(paths.TENDERS_DIR, "reports")
        self.live_dir = os.path.join(self.reports_dir, "live")
        self.daily_dir = os.path.join(self.reports_dir, "daily_exports")

        os.makedirs(self.live_dir, exist_ok=True)
        os.makedirs(self.daily_dir, exist_ok=True)

        self.live_excel_path = os.path.join(self.live_dir, "live_working_excel.xlsx")
        self.state_file = os.path.join(self.live_dir, ".live_session.json")

        self.is_active: bool = False
        self.status: str = "idle"  # "idle" | "active" | "saved" | "removed"
        self.created_at: Optional[float] = None
        self.last_updated_at: Optional[float] = None
        self.update_count: int = 0
        self.tender_bids: List[str] = []
        self.last_saved_filename: Optional[str] = None
        self.last_message: str = "Ready"

        # Restore any ongoing session from disk or clean up stale
        self._load_state()

        # Start watcher thread
        self._stop_event = threading.Event()
        self._watcher_thread = threading.Thread(target=self._watcher_loop, daemon=True)
        self._watcher_thread.start()

    def _save_state(self):
        try:
            state = {
                "is_active": self.is_active,
                "status": self.status,
                "created_at": self.created_at,
                "last_updated_at": self.last_updated_at,
                "update_count": self.update_count,
                "tender_bids": self.tender_bids,
                "last_saved_filename": self.last_saved_filename,
                "last_message": self.last_message,
            }
            with open(self.state_file, "w", encoding="utf-8") as f:
                json.dump(state, f, indent=2)
        except Exception as e:
            logger.warning("Could not persist live excel state: %s", e)

    def _load_state(self):
        if not os.path.exists(self.state_file):
            return
        try:
            with open(self.state_file, "r", encoding="utf-8") as f:
                state = json.load(f)
            if state.get("is_active"):
                last_up = state.get("last_updated_at", 0)
                now = time.time()
                if (now - last_up) >= self.timeout_seconds:
                    # Stale session on startup -> finalize or remove
                    if state.get("update_count", 0) == 0 or not state.get("tender_bids"):
                        self.is_active = False
                        self.status = "removed"
                        self.last_message = "Previous session expired while offline (0 updates, removed)."
                        if os.path.exists(self.live_excel_path):
                            try:
                                os.remove(self.live_excel_path)
                            except Exception:
                                pass
                    else:
                        self.tender_bids = state.get("tender_bids", [])
                        self._finalize_and_save("Auto-saved session from previous run.")
                else:
                    self.is_active = True
                    self.status = "active"
                    self.created_at = state.get("created_at")
                    self.last_updated_at = last_up
                    self.update_count = state.get("update_count", 0)
                    self.tender_bids = state.get("tender_bids", [])
                    self.last_saved_filename = state.get("last_saved_filename")
                    self.last_message = "Resumed active session."
            else:
                self.status = state.get("status", "idle")
                self.last_saved_filename = state.get("last_saved_filename")
        except Exception as e:
            logger.warning("Failed loading live excel state: %s", e)

    def _watcher_loop(self):
        """Background checker for inactivity timeout."""
        while not self._stop_event.is_set():
            sleep_time = min(0.2, max(0.05, self.timeout_seconds / 4))
            time.sleep(sleep_time)
            try:
                with self.lock:
                    if not self.is_active or self.last_updated_at is None:
                        continue
                    idle_seconds = time.time() - self.last_updated_at
                    if idle_seconds >= self.timeout_seconds:
                        if self.update_count == 0 or not self.tender_bids:
                            self._discard_session("Removed: 10 minutes elapsed with no updates.")
                        else:
                            self._finalize_and_save("Closed & saved: 10 minutes elapsed since last update.")
            except Exception as e:
                logger.error("Error in live excel watcher: %s", e)

    def _get_next_daily_filename(self) -> str:
        """Find the next sequential filename for today: YYYY-MM-DD_<N>.xlsx."""
        today = datetime.date.today().isoformat()
        pattern = os.path.join(self.daily_dir, f"{today}_*.xlsx")
        existing = glob.glob(pattern)
        max_idx = 0
        rx = re.compile(rf"^{re.escape(today)}_(\d+)\.xlsx$", re.IGNORECASE)
        for p in existing:
            basename = os.path.basename(p)
            m = rx.match(basename)
            if m:
                try:
                    idx = int(m.group(1))
                    if idx > max_idx:
                        max_idx = idx
                except ValueError:
                    pass
        next_idx = max_idx + 1
        return f"{today}_{next_idx}.xlsx"

    def _resolve_tenders(self, bids: List[str]) -> List[Dict[str, Any]]:
        """Load tender metadata dictionaries corresponding to bid numbers."""
        try:
            from gemsentry.storage import load_existing_metadata
            all_tenders = load_existing_metadata()
            result = []
            for b in bids:
                if b in all_tenders:
                    result.append(all_tenders[b])
                else:
                    result.append({"bid_no": b, "title": "Unknown Bid", "status": "Pending Review"})
            return result
        except Exception as e:
            logger.error("Error resolving tender metadata for excel: %s", e)
            return [{"bid_no": b} for b in bids]

    def _refresh_live_workbook(self):
        """Builds or rebuilds the live working Excel file on disk."""
        tenders = self._resolve_tenders(self.tender_bids)
        build_curated_workbook(tenders, self.live_excel_path)
        logger.info("Live Excel refreshed with %d tenders.", len(tenders))

    def _discard_session(self, reason: str = "Session discarded"):
        """Discard the session and delete the temporary live excel."""
        self.is_active = False
        self.status = "removed"
        self.last_message = reason
        self.tender_bids = []
        self.update_count = 0
        self.created_at = None
        self.last_updated_at = None
        if os.path.exists(self.live_excel_path):
            try:
                os.remove(self.live_excel_path)
            except Exception as e:
                logger.warning("Could not delete live excel file: %s", e)
        self._save_state()
        logger.info("Live Excel session discarded: %s", reason)

    def _finalize_and_save(self, reason: str = "Session closed and saved") -> Optional[str]:
        """Save the working Excel permanently with sequential daily numbering."""
        if not self.tender_bids:
            self._discard_session("No tenders in session upon close.")
            return None

        filename = self._get_next_daily_filename()
        target_path = os.path.join(self.daily_dir, filename)

        try:
            # Build final workbook directly at target
            tenders = self._resolve_tenders(self.tender_bids)
            build_curated_workbook(tenders, target_path)

            # Clean up live working file
            if os.path.exists(self.live_excel_path):
                try:
                    os.remove(self.live_excel_path)
                except Exception:
                    pass

            self.is_active = False
            self.status = "saved"
            self.last_saved_filename = filename
            self.last_message = f"{reason} Saved as {filename} ({len(tenders)} tenders)."
            saved_bids_count = len(self.tender_bids)
            self.tender_bids = []
            self.update_count = 0
            self.created_at = None
            self.last_updated_at = None
            self._save_state()

            logger.info("Live Excel saved permanently as %s (%d tenders).", filename, saved_bids_count)
            return filename
        except Exception as e:
            logger.error("Failed to save final live excel: %s", e)
            self.last_message = f"Error saving live excel: {e}"
            self._save_state()
            return None

    # -------------------------------------------------------------------------
    # Public API
    # -------------------------------------------------------------------------

    def on_scrape_completed(self, new_tenders: Optional[List[Dict[str, Any]]] = None):
        """
        Called after each scrape completes.
        Stores an Excel session with a 10-minute timer.
        If no updates occur within 10 minutes, it is removed.
        """
        with self.lock:
            # If previous session had updates, finalize and save it first
            if self.is_active and self.update_count > 0 and self.tender_bids:
                self._finalize_and_save("Saved previous session prior to new scrape.")
            elif self.is_active:
                self._discard_session("Discarded previous untouched session prior to new scrape.")

            now = time.time()
            self.is_active = True
            self.status = "active"
            self.created_at = now
            self.last_updated_at = now
            self.update_count = 0
            self.tender_bids = []
            self.last_message = "Live Excel session started after scrape. 10m inactivity window open."

            # Generate the initial base workbook
            self._refresh_live_workbook()
            self._save_state()
            logger.info("Live Excel session started after scrape.")

    def toggle_tender(self, bid_no: str) -> Dict[str, Any]:
        """
        Toggle inclusion of a tender in the active Excel.
        If the 10 minutes have expired or session is idle, forms a NEW session.
        Resets the 10-minute timer!
        """
        with self.lock:
            now = time.time()
            # If inactive or expired, start a brand new session
            if not self.is_active:
                self.is_active = True
                self.status = "active"
                self.created_at = now
                self.tender_bids = []
                self.update_count = 0
                logger.info("Forming new Live Excel session on user click for bid %s", bid_no)

            action = ""
            if bid_no in self.tender_bids:
                self.tender_bids.remove(bid_no)
                action = "removed"
            else:
                self.tender_bids.append(bid_no)
                action = "added"

            self.update_count += 1
            self.last_updated_at = now  # Timer reset!
            self.last_message = f"Tender {bid_no} {action}. 10m countdown reset."

            self._refresh_live_workbook()
            self._save_state()

            return {
                "action": action,
                "bid_no": bid_no,
                "status": self.get_status(touch=False),
            }

    def add_tender(self, bid_no: str) -> Dict[str, Any]:
        """Explicitly add a tender to active Excel (starts new session if idle)."""
        with self.lock:
            now = time.time()
            if not self.is_active:
                self.is_active = True
                self.status = "active"
                self.created_at = now
                self.tender_bids = []
                self.update_count = 0

            if bid_no not in self.tender_bids:
                self.tender_bids.append(bid_no)
                self.update_count += 1
                self.last_updated_at = now  # Timer reset!
                self.last_message = f"Tender {bid_no} added. 10m countdown reset."
                self._refresh_live_workbook()
                self._save_state()

            return self.get_status(touch=False)

    def remove_tender(self, bid_no: str) -> Dict[str, Any]:
        """Explicitly remove a tender from active Excel."""
        with self.lock:
            if not self.is_active:
                return self.get_status(touch=False)

            if bid_no in self.tender_bids:
                self.tender_bids.remove(bid_no)
                self.update_count += 1
                self.last_updated_at = time.time()  # Timer reset!
                self.last_message = f"Tender {bid_no} removed. 10m countdown reset."
                self._refresh_live_workbook()
                self._save_state()

            return self.get_status(touch=False)

    def add_batch(self, bid_nos: List[str]) -> Dict[str, Any]:
        """Add multiple tenders at once (e.g. 'Add All Proceed'). Resets 10m timer."""
        with self.lock:
            now = time.time()
            if not self.is_active:
                self.is_active = True
                self.status = "active"
                self.created_at = now
                self.tender_bids = []
                self.update_count = 0

            added = 0
            for b in bid_nos:
                if b and b not in self.tender_bids:
                    self.tender_bids.append(b)
                    added += 1

            if added > 0:
                self.update_count += 1
                self.last_updated_at = now  # Timer reset!
                self.last_message = f"Added {added} tenders in batch. 10m countdown reset."
                self._refresh_live_workbook()
                self._save_state()

            return self.get_status(touch=False)

    def manual_close(self, save: bool = True) -> Dict[str, Any]:
        """Manually save and close or discard the session without waiting 10 minutes."""
        with self.lock:
            if not self.is_active:
                return {"error": "No active Live Excel session."}

            if save:
                filename = self._finalize_and_save("Manually closed and saved.")
                return {"message": "Session saved.", "filename": filename, "status": self.get_status(touch=False)}
            else:
                self._discard_session("Manually discarded by user.")
                return {"message": "Session discarded.", "status": self.get_status(touch=False)}

    def list_today_saved_files(self) -> List[Dict[str, Any]]:
        """List today's saved Excel files (and recent daily files)."""
        today = datetime.date.today().isoformat()
        pattern = os.path.join(self.daily_dir, "*.xlsx")
        files = []
        for p in sorted(glob.glob(pattern), reverse=True):
            name = os.path.basename(p)
            try:
                stat = os.stat(p)
                mtime = datetime.datetime.fromtimestamp(stat.st_mtime).strftime("%I:%M %p, %d %b")
                is_today = name.startswith(today)
                files.append({
                    "filename": name,
                    "size_bytes": stat.st_size,
                    "modified": mtime,
                    "is_today": is_today,
                    "url": f"/api/live-excel/download/saved/{urllib.parse.quote(name)}",
                })
            except Exception:
                pass
        return files

    def get_status(self, touch: bool = False) -> Dict[str, Any]:
        """
        Read status of the Live Excel manager.
        READ DOES NOT COUNT -> touch defaults to False, so timer is not reset!
        """
        with self.lock:
            seconds_remaining = 0
            if self.is_active and self.last_updated_at is not None:
                idle = time.time() - self.last_updated_at
                seconds_remaining = max(0, int(self.timeout_seconds - idle))

            return {
                "is_active": self.is_active,
                "status": self.status,
                "seconds_remaining": seconds_remaining,
                "timeout_seconds": self.timeout_seconds,
                "update_count": self.update_count,
                "tender_bids": list(self.tender_bids),
                "tender_count": len(self.tender_bids),
                "created_at": (
                    datetime.datetime.fromtimestamp(self.created_at).isoformat()
                    if self.created_at else None
                ),
                "last_updated_at": (
                    datetime.datetime.fromtimestamp(self.last_updated_at).isoformat()
                    if self.last_updated_at else None
                ),
                "last_saved_filename": self.last_saved_filename,
                "last_message": self.last_message,
                "has_live_file": os.path.exists(self.live_excel_path),
                "today_saved_files": self.list_today_saved_files(),
            }


# Singleton instance for the server process
live_excel_manager = LiveExcelManager()
